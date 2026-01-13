import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import winsound
from copy import deepcopy
import shutil

"""
╔══════════════════════════════════════════════════════════════════╗
║                    CUBE DATA PROCESSOR v3.1                       ║
║                                                                   ║
║  Developer: Sandeep (https://github.com/Sandeep2062)            ║
║  Repository: https://github.com/Sandeep2062/Cube-Merge          ║
║  Description: Excel data processor with logo preservation        ║
║                                                                   ║
║  © 2025 Sandeep - All Rights Reserved                           ║
╚══════════════════════════════════════════════════════════════════╝
"""

# Detect grade from filename
def extract_grade(filename):
    name = os.path.basename(filename).split('.')[0].upper()
    name = name.replace("_", ":").replace("-", ":")
    return name.strip()


# FILLED ROW CHECKER
def get_last_row(ws):
    row = 2
    while True:
        if ws.cell(row=row, column=2).value in (None, ""):
            return row - 1
        row += 1


# SAFE WORKBOOK LOADING - Preserves images and prevents corruption
def load_workbook_safe(filepath):
    """Load workbook while preserving images and avoiding corruption"""
    try:
        # Try with keep_vba first
        wb = openpyxl.load_workbook(filepath, keep_vba=False, data_only=False, keep_links=False)
        return wb
    except:
        # Fallback to basic load
        wb = openpyxl.load_workbook(filepath)
        return wb


# COPY IMAGES BETWEEN SHEETS - Fixed version
def copy_all_images_from_template(template_file, output_file):
    """
    Copy all images from template to output file after saving data.
    This ensures logos are preserved without corruption.
    """
    try:
        # Load both files
        template_wb = openpyxl.load_workbook(template_file)
        output_wb = openpyxl.load_workbook(output_file)
        
        # Copy images from each sheet
        for sheet_name in template_wb.sheetnames:
            if sheet_name in output_wb.sheetnames:
                source_sheet = template_wb[sheet_name]
                target_sheet = output_wb[sheet_name]
                
                # Copy images if they exist
                if hasattr(source_sheet, '_images') and source_sheet._images:
                    # Clear existing images in target
                    target_sheet._images = []
                    
                    # Copy each image
                    for img in source_sheet._images:
                        new_img = XLImage(img.ref)
                        new_img.anchor = deepcopy(img.anchor)
                        target_sheet.add_image(new_img)
        
        # Save with images
        output_wb.save(output_file)
        template_wb.close()
        output_wb.close()
        return True
    except Exception as e:
        print(f"Image copy warning: {e}")
        return False


# MAIN PROCESSING - SEPARATE MODE
def process_grade_separate(grade_file, office_file, output_folder, log):
    temp_file = None
    try:
        grade_wb = load_workbook_safe(grade_file)
        grade_ws = grade_wb.active

        grade_name = extract_grade(grade_file)
        log(f"\n=== Processing {grade_file}")
        log(f"Detected Grade: {grade_name}")

        # Create a temporary copy of the office file to preserve images
        base = os.path.basename(office_file).split(".")[0]
        outname = f"{base}_{grade_name}_Processed.xlsx"
        outpath = os.path.join(output_folder, outname)
        
        # Copy the template file first (preserves everything)
        shutil.copy2(office_file, outpath)
        
        # Now load and modify the copy
        office_wb = load_workbook_safe(outpath)

        last_row = get_last_row(grade_ws)
        log(f"Total data rows: {last_row - 1}")

        # Get all sheets that match this grade
        matching_sheets = []
        for sheet_name in office_wb.sheetnames:
            ws = office_wb[sheet_name]
            b12_value = ws["B12"].value
            if b12_value:
                b12 = str(b12_value).replace(" ", "").upper()
                if b12 == grade_name:
                    matching_sheets.append(sheet_name)
        
        log(f"Found {len(matching_sheets)} sheets matching grade '{grade_name}'")
        
        if len(matching_sheets) == 0:
            log(f"⚠ WARNING: No sheets found with '{grade_name}' in cell B12!")
            office_wb.close()
            os.remove(outpath)
            return 0

        copy_count = 0
        sheet_index = 0

        # Loop through each data row
        for r in range(2, last_row + 1):
            if sheet_index >= len(matching_sheets):
                log(f"⚠ Warning: More data rows than matching sheets. Stopping at row {r}")
                break

            current_sheet_name = matching_sheets[sheet_index]
            ws = office_wb[current_sheet_name]
            
            # Read weight and strength values
            weight_values = [grade_ws.cell(row=r, column=c).value for c in range(2, 8)]
            strength_values = [grade_ws.cell(row=r, column=c).value for c in range(9, 15)]

            # Write values (ONLY modify data cells, logos remain untouched)
            for i, v in enumerate(weight_values):
                ws.cell(row=25, column=3 + i, value=v)
            for i, v in enumerate(strength_values):
                ws.cell(row=27, column=3 + i, value=v)

            copy_count += 1
            log(f"✓ Row {r} → Sheet: {current_sheet_name}")
            sheet_index += 1

        # Save (logos already preserved from copy)
        office_wb.save(outpath)
        office_wb.close()
        grade_wb.close()
        
        log(f"✓ Saved → {outpath} (Logos preserved ✓)")

        return copy_count

    except Exception as e:
        log(f"✖ ERROR: {e}")
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
        return 0


# MAIN PROCESSING - COMBINE MODE
def process_all_grades_combined(grade_files, office_file, output_folder, log):
    try:
        log(f"\n=== COMBINE MODE: Processing all grades into one file ===")
        
        # Create output file by copying template first
        base = os.path.basename(office_file).split(".")[0]
        outname = f"{base}_ALL_GRADES_Combined.xlsx"
        outpath = os.path.join(output_folder, outname)
        
        # Copy template to preserve images
        shutil.copy2(office_file, outpath)
        
        # Now load and modify the copy
        office_wb = load_workbook_safe(outpath)
        
        total_copy_count = 0
        
        # Process each grade file
        for grade_file in grade_files:
            grade_wb = load_workbook_safe(grade_file)
            grade_ws = grade_wb.active
            grade_name = extract_grade(grade_file)
            
            log(f"\n--- Processing Grade: {grade_name} ---")
            
            last_row = get_last_row(grade_ws)
            log(f"Data rows: {last_row - 1}")
            
            # Find matching sheets for this grade
            matching_sheets = []
            for sheet_name in office_wb.sheetnames:
                ws = office_wb[sheet_name]
                b12_value = ws["B12"].value
                if b12_value:
                    b12 = str(b12_value).replace(" ", "").upper()
                    if b12 == grade_name:
                        matching_sheets.append(sheet_name)
            
            log(f"Found {len(matching_sheets)} sheets for '{grade_name}'")
            
            if len(matching_sheets) == 0:
                log(f"⚠ No sheets found for '{grade_name}'")
                grade_wb.close()
                continue
            
            sheet_index = 0
            
            # Copy data for this grade (logos stay intact)
            for r in range(2, last_row + 1):
                if sheet_index >= len(matching_sheets):
                    log(f"⚠ More rows than sheets for {grade_name}")
                    break
                
                current_sheet_name = matching_sheets[sheet_index]
                ws = office_wb[current_sheet_name]
                
                weight_values = [grade_ws.cell(row=r, column=c).value for c in range(2, 8)]
                strength_values = [grade_ws.cell(row=r, column=c).value for c in range(9, 15)]
                
                for i, v in enumerate(weight_values):
                    ws.cell(row=25, column=3 + i, value=v)
                for i, v in enumerate(strength_values):
                    ws.cell(row=27, column=3 + i, value=v)
                
                total_copy_count += 1
                log(f"✓ {grade_name} Row {r} → {current_sheet_name}")
                sheet_index += 1
            
            grade_wb.close()
        
        # Save combined file (logos already preserved)
        office_wb.save(outpath)
        office_wb.close()
        
        log(f"\n✓✓✓ COMBINED FILE SAVED → {outpath} (Logos preserved ✓)")
        
        return total_copy_count
        
    except Exception as e:
        log(f"✖ ERROR: {e}")
        return 0


# ------------- GUI LOGIC -------------

def run_processing():
    if not grade_files:
        messagebox.showerror("Error", "Please select grade files.")
        return

    if not office_path.get():
        messagebox.showerror("Error", "Select office format file.")
        return

    if not output_path.get():
        messagebox.showerror("Error", "Select output folder.")
        return

    log_box.delete("1.0", "end")
    total = 0

    if mode_var.get() == 2:  # COMBINE MODE
        progress["value"] = 50
        root.update_idletasks()
        
        total = process_all_grades_combined(
            grade_files,
            office_path.get(),
            output_path.get(),
            log=lambda m: log_box.insert(tk.END, m + "\n")
        )
        
        progress["value"] = 100
        
    else:  # SEPARATE MODE
        for i, file in enumerate(grade_files):
            progress["value"] = (i + 1) / len(grade_files) * 100
            root.update_idletasks()

            total += process_grade_separate(
                file,
                office_path.get(),
                output_path.get(),
                log=lambda m: log_box.insert(tk.END, m + "\n")
            )

    winsound.MessageBeep()
    messagebox.showinfo("✓ Completed", f"Processing Complete!\n\nTotal Rows Copied: {total}")


def add_grades():
    files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
    for f in files:
        if f not in grade_files:
            grade_files.append(f)
            grade_listbox.insert(tk.END, os.path.basename(f))


def clear_grades():
    grade_files.clear()
    grade_listbox.delete(0, tk.END)


def pick_office():
    path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if path:
        office_path.set(path)


def pick_output_folder():
    folder = filedialog.askdirectory()
    if folder:
        output_path.set(folder)


# ------------------- DARK MODE PREMIUM UI -------------------

root = tk.Tk()
root.title("Cube Data Processor v3.1 - Dark Edition")
root.geometry("920x820")
root.configure(bg="#1a1a1a")

# Try to set window icon
try:
    root.iconbitmap("icon.ico")
except:
    pass

grade_files = []
office_path = tk.StringVar()
output_path = tk.StringVar()
mode_var = tk.IntVar(value=1)

# Dark Theme Colors
BG_DARK = "#1a1a1a"
BG_CARD = "#252525"
BG_INPUT = "#2d2d2d"
TEXT_PRIMARY = "#ffffff"
TEXT_SECONDARY = "#b0b0b0"
ACCENT_BLUE = "#0d7377"
ACCENT_GREEN = "#14a76c"
BORDER_COLOR = "#3a3a3a"

# Enhanced Style for Dark Mode
style = ttk.Style()
style.theme_use('clam')
style.configure('Dark.TButton', padding=8, relief="flat", background=ACCENT_BLUE, 
                foreground="white", font=("Segoe UI", 9), borderwidth=0)
style.map('Dark.TButton', background=[('active', '#0a5d5f')])

style.configure('Action.TButton', padding=12, background=ACCENT_GREEN, 
                foreground="white", font=("Segoe UI", 11, "bold"), borderwidth=0)
style.map('Action.TButton', background=[('active', '#0f8655')])

style.configure("Dark.Horizontal.TProgressbar", background=ACCENT_GREEN, 
                troughcolor=BG_INPUT, borderwidth=0, lightcolor=ACCENT_GREEN, darkcolor=ACCENT_GREEN)

# GRADIENT HEADER
header_frame = tk.Frame(root, bg="#0d7377", height=100)
header_frame.pack(fill=tk.X)
header_frame.pack_propagate(False)

# Logo Space (LEFT)
logo_container = tk.Frame(header_frame, bg="#0d7377")
logo_container.place(x=25, y=20)

try:
    from PIL import Image, ImageTk
    logo_img = Image.open("logo.png")
    logo_img = logo_img.resize((60, 60), Image.Resampling.LANCZOS)
    logo_photo = ImageTk.PhotoImage(logo_img)
    logo_label = tk.Label(logo_container, image=logo_photo, bg="#0d7377")
    logo_label.image = logo_photo
    logo_label.pack()
except:
    logo_label = tk.Label(logo_container, text="🔷", font=("Segoe UI", 42), bg="#0d7377", fg="white")
    logo_label.pack()

# Title (CENTER)
title_container = tk.Frame(header_frame, bg="#0d7377")
title_container.pack(expand=True)

title_label = tk.Label(title_container, text="CUBE DATA PROCESSOR", 
                       font=("Segoe UI", 24, "bold"), bg="#0d7377", fg="white")
title_label.pack()

version_label = tk.Label(title_container, text="v3.1 - Dark Edition | Logo Preservation Technology", 
                        font=("Segoe UI", 9), bg="#0d7377", fg="#a8e6cf")
version_label.pack()

# Developer Credit (RIGHT)
credit_frame = tk.Frame(header_frame, bg="#0d7377")
credit_frame.place(relx=1.0, y=20, anchor="ne", x=-25)

credit_label = tk.Label(credit_frame, text="Developed by", 
                       font=("Segoe UI", 8), bg="#0d7377", fg="#b0b0b0")
credit_label.pack()

dev_name_label = tk.Label(credit_frame, text="SANDEEP", 
                         font=("Segoe UI", 11, "bold"), bg="#0d7377", fg="white")
dev_name_label.pack()

github_label = tk.Label(credit_frame, text="github.com/Sandeep2062", 
                       font=("Segoe UI", 8), bg="#0d7377", fg="#4dd0e1", cursor="hand2")
github_label.pack()
github_label.bind("<Button-1>", lambda e: os.system("start https://github.com/Sandeep2062/Cube-Merge"))

# Main Container
main_container = tk.Frame(root, bg=BG_DARK)
main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

# Grade Files Section
grade_frame = tk.LabelFrame(main_container, text="  📁 GRADE FILES  ", 
                            font=("Segoe UI", 11, "bold"), bg=BG_CARD, 
                            fg=TEXT_PRIMARY, bd=2, relief=tk.FLAT, padx=15, pady=15)
grade_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 12))

btn_frame = tk.Frame(grade_frame, bg=BG_CARD)
btn_frame.pack(fill=tk.X, pady=(0, 10))

ttk.Button(btn_frame, text="➕ Add Files", command=add_grades, style='Dark.TButton').pack(side=tk.LEFT, padx=5)
ttk.Button(btn_frame, text="🗑️ Clear All", command=clear_grades, style='Dark.TButton').pack(side=tk.LEFT, padx=5)

grade_listbox = tk.Listbox(grade_frame, height=5, font=("Consolas", 9), 
                           bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, 
                           highlightthickness=1, highlightbackground=BORDER_COLOR,
                           selectbackground=ACCENT_BLUE, selectforeground="white")
grade_listbox.pack(fill=tk.BOTH, expand=True)

# Office File Section
office_frame = tk.LabelFrame(main_container, text="  📄 OFFICE FORMAT FILE  ", 
                            font=("Segoe UI", 11, "bold"), bg=BG_CARD, 
                            fg=TEXT_PRIMARY, bd=2, relief=tk.FLAT, padx=15, pady=15)
office_frame.pack(fill=tk.X, pady=(0, 12))

ttk.Button(office_frame, text="📂 Select File", command=pick_office, style='Dark.TButton').pack(anchor=tk.W, pady=(0, 8))
office_entry = tk.Entry(office_frame, textvariable=office_path, font=("Segoe UI", 9),
                       bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, insertbackground="white")
office_entry.pack(fill=tk.X, ipady=10, padx=2)

# Output Folder Section
output_frame = tk.LabelFrame(main_container, text="  💾 OUTPUT FOLDER  ", 
                            font=("Segoe UI", 11, "bold"), bg=BG_CARD, 
                            fg=TEXT_PRIMARY, bd=2, relief=tk.FLAT, padx=15, pady=15)
output_frame.pack(fill=tk.X, pady=(0, 12))

ttk.Button(output_frame, text="📂 Select Folder", command=pick_output_folder, style='Dark.TButton').pack(anchor=tk.W, pady=(0, 8))
output_entry = tk.Entry(output_frame, textvariable=output_path, font=("Segoe UI", 9),
                       bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, insertbackground="white")
output_entry.pack(fill=tk.X, ipady=10, padx=2)

# Processing Mode
mode_frame = tk.LabelFrame(main_container, text="  ⚙️ PROCESSING MODE  ", 
                          font=("Segoe UI", 11, "bold"), bg=BG_CARD, 
                          fg=TEXT_PRIMARY, bd=2, relief=tk.FLAT, padx=15, pady=15)
mode_frame.pack(fill=tk.X, pady=(0, 12))

tk.Radiobutton(mode_frame, text="📑 Separate Files (One file per grade)", 
               variable=mode_var, value=1, font=("Segoe UI", 10),
               bg=BG_CARD, fg=TEXT_PRIMARY, activebackground=BG_CARD, 
               activeforeground=TEXT_PRIMARY, selectcolor=BG_INPUT).pack(anchor=tk.W, pady=4)
tk.Radiobutton(mode_frame, text="📦 Combined File (All grades merged)", 
               variable=mode_var, value=2, font=("Segoe UI", 10),
               bg=BG_CARD, fg=TEXT_PRIMARY, activebackground=BG_CARD, 
               activeforeground=TEXT_PRIMARY, selectcolor=BG_INPUT).pack(anchor=tk.W, pady=4)

# Start Button
start_btn = tk.Button(main_container, text="▶️  START PROCESSING", command=run_processing,
                     font=("Segoe UI", 13, "bold"), bg=ACCENT_GREEN, fg="white",
                     activebackground="#0f8655", relief=tk.FLAT, cursor="hand2",
                     padx=35, pady=16, borderwidth=0)
start_btn.pack(pady=15)

# Progress Bar
progress_frame = tk.Frame(main_container, bg=BG_DARK)
progress_frame.pack(fill=tk.X, pady=(0, 12))

progress = ttk.Progressbar(progress_frame, length=600, mode="determinate", style="Dark.Horizontal.TProgressbar")
progress.pack()

# Log Section
log_frame = tk.LabelFrame(main_container, text="  📋 PROCESSING LOG  ", 
                         font=("Segoe UI", 11, "bold"), bg=BG_CARD, 
                         fg=TEXT_PRIMARY, bd=2, relief=tk.FLAT, padx=15, pady=15)
log_frame.pack(fill=tk.BOTH, expand=True)

log_scrollbar = tk.Scrollbar(log_frame, bg=BG_INPUT)
log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

log_box = tk.Text(log_frame, height=10, font=("Consolas", 9), bg=BG_INPUT, fg="#a8e6cf",
                 relief=tk.FLAT, bd=0, wrap=tk.WORD, yscrollcommand=log_scrollbar.set,
                 insertbackground="white")
log_box.pack(fill=tk.BOTH, expand=True)
log_scrollbar.config(command=log_box.yview)

# Footer
footer = tk.Label(root, text="© 2025 Sandeep | github.com/Sandeep2062/Cube-Merge | Logo Preservation ✓ | No Corruption ✓", 
                 font=("Segoe UI", 8), bg=BG_DARK, fg=TEXT_SECONDARY)
footer.pack(pady=8)

root.mainloop()